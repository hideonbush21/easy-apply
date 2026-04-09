"""
从 Excel 导入院校专业数据到 schools + programs 表。
用法：python seed_data/import_programs.py <excel_path>
"""

import sys
import re
import os
import openpyxl
from datetime import datetime

# 列索引（0-based）
COL = {
    'project_code':       0,   # Project Internal ID
    'school_code':        1,   # School Internal ID
    'name_cn':            2,   # 学校名称
    'name_en':            3,   # School Name (EN)
    'ranking_2026':       4,   # QS Rank 2026
    'prog_name_cn':       5,   # 项目名称
    'prog_name_en':       6,   # Major (EN)
    'major_category':     7,   # 专业大类
    'department':         8,   # 学院
    'intake_month':       9,   # 入学时间
    'duration':          10,   # 学制
    'tuition':           11,   # Tuition
    'tuition_cny':       12,   # 学费 人民币
    'description':       13,   # 专业介绍
    'requirements':      14,   # 录取要求
    'ielts_raw':         16,   # 雅思（格式：6.5(6.0)）
    'toefl_raw':         18,   # 托福（格式：92(20)）
    'pte_raw':           20,   # PTE（格式：62(56)）
    'deadline_26fall_start': 21,
    'deadline_26fall':   22,   # 26fall截止时间
    'plan_26fall':       23,   # 26fall申请时间计划
    'deadline_25fall_start': 24,
    'deadline_25fall':   25,   # 25fall截止时间
    'plan_25fall':       26,   # 25fall申请时间计划
    'program_url':       27,   # 项目链接
    'country_en':        28,   # Country
    'country_cn':        29,   # 国家|地区
    'ranking_2024':      30,   # QS Rank 2024
    'ranking_2025':      31,   # QS Rank 2025
}


def parse_lang_req(raw):
    """解析 '6.5(6.0)' → {"total": 6.5, "sub": 6.0}，空值返回 None"""
    if not raw:
        return None
    raw = str(raw).strip()
    m = re.match(r'([\d.]+)\(([\d.]+)\)', raw)
    if m:
        return {"total": float(m.group(1)), "sub": float(m.group(2))}
    # 只有总分没有小分
    m2 = re.match(r'([\d.]+)', raw)
    if m2:
        return {"total": float(m2.group(1))}
    return None


def parse_int(val):
    if val is None:
        return None
    try:
        return int(str(val).strip())
    except (ValueError, TypeError):
        return None


def parse_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    try:
        return datetime.strptime(str(val).strip(), '%Y-%m-%d').date()
    except (ValueError, TypeError):
        return None


def extract_url(val):
    """从 HYPERLINK 公式或普通字符串中提取 URL"""
    if not val:
        return None
    s = str(val)
    m = re.search(r'HYPERLINK\("([^"]+)"', s)
    if m:
        return m.group(1)
    return s.strip() if s.startswith('http') else None


def row_to_dict(row):
    def v(col_key):
        idx = COL[col_key]
        return row[idx] if idx < len(row) else None

    return {
        'school': {
            'school_code': str(v('school_code')).strip() if v('school_code') else None,
            'name_cn':     str(v('name_cn')).strip() if v('name_cn') else None,
            'name':        str(v('name_en')).strip() if v('name_en') else None,
            'country':     str(v('country_cn')).strip() if v('country_cn') else (
                           str(v('country_en')).strip() if v('country_en') else '未知'),
            'ranking':     parse_int(v('ranking_2026')),
            'ranking_2024': parse_int(v('ranking_2024')),
            'ranking_2025': parse_int(v('ranking_2025')),
        },
        'program': {
            'project_code':      str(v('project_code')).strip() if v('project_code') else None,
            'name_cn':           str(v('prog_name_cn')).strip() if v('prog_name_cn') else None,
            'name_en':           str(v('prog_name_en')).strip() if v('prog_name_en') else None,
            'major_category':    str(v('major_category')).strip() if v('major_category') else None,
            'department':        str(v('department')).strip() if v('department') else None,
            'intake_month':      str(v('intake_month')).strip() if v('intake_month') else None,
            'duration':          str(v('duration')).strip() if v('duration') else None,
            'tuition':           str(v('tuition')).strip() if v('tuition') else None,
            'tuition_cny':       parse_int(v('tuition_cny')),
            'description':       str(v('description')).strip() if v('description') else None,
            'requirements':      str(v('requirements')).strip() if v('requirements') else None,
            'ielts_requirement': parse_lang_req(v('ielts_raw')),
            'toefl_requirement': parse_lang_req(v('toefl_raw')),
            'pte_requirement':   parse_lang_req(v('pte_raw')),
            'deadline_26fall':   parse_date(v('deadline_26fall')),
            'deadline_25fall':   parse_date(v('deadline_25fall')),
            'application_plan':  {
                '26fall': str(v('plan_26fall')).strip() if v('plan_26fall') else None,
                '25fall': str(v('plan_25fall')).strip() if v('plan_25fall') else None,
            },
            'program_url':       extract_url(v('program_url')),
        }
    }


def import_excel(excel_path):
    print(f"读取文件：{excel_path}")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active
    total_rows = ws.max_row - 1  # 减去表头
    print(f"共 {total_rows} 条记录，开始导入...")

    # 延迟导入 Flask 应用（需要在 backend 目录下运行）
    sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    from app import create_app
    from app.extensions import db
    from app.models.school import School
    from app.models.program import Program

    app = create_app()
    with app.app_context():
        # 先建表（如果不存在）
        db.create_all()

        school_cache = {}   # school_code → School 对象
        prog_inserted = 0
        school_inserted = 0
        school_updated = 0

        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
            if not any(row):
                continue
            try:
                data = row_to_dict(row)
            except Exception as e:
                print(f"  [跳过] 第 {i+1} 行解析失败：{e}")
                continue

            sd = data['school']
            pd = data['program']
            school_code = sd['school_code']

            # 1. 写入/更新 school
            if school_code not in school_cache:
                school = School.query.filter_by(school_code=school_code).first()
                if not school:
                    school = School(
                        school_code=school_code,
                        name=sd['name'],
                        name_cn=sd['name_cn'],
                        country=sd['country'],
                        ranking=sd['ranking'],
                        ranking_2024=sd['ranking_2024'],
                        ranking_2025=sd['ranking_2025'],
                    )
                    db.session.add(school)
                    db.session.flush()  # 获取 id
                    school_inserted += 1
                else:
                    # 更新排名字段
                    school.ranking = sd['ranking']
                    school.ranking_2024 = sd['ranking_2024']
                    school.ranking_2025 = sd['ranking_2025']
                    school_updated += 1
                school_cache[school_code] = school
            else:
                school = school_cache[school_code]

            # 2. 写入 program（以 project_code 去重）
            if pd['project_code']:
                exists = Program.query.filter_by(project_code=pd['project_code']).first()
                if exists:
                    continue

            program = Program(
                school_id=school.id,
                **pd
            )
            db.session.add(program)
            prog_inserted += 1

            if i % 200 == 0:
                db.session.commit()
                print(f"  进度：{i}/{total_rows}")

        db.session.commit()
        print(f"\n✅ 导入完成：")
        print(f"   学校新增 {school_inserted} 所，更新 {school_updated} 所")
        print(f"   专业新增 {prog_inserted} 条")


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("用法：python seed_data/import_programs.py <excel_path>")
        sys.exit(1)
    import_excel(sys.argv[1])
