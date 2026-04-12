"""
邮件处理模块

负责：
1. 附件文本提取（PDF / Word / Excel / 图片占位）
2. 调用 Kimi LLM 分析邮件内容，返回结构化 JSON
   - 使用与 backend/app/services/ai_service.py 完全一致的 Kimi 调用模式
"""

import io
import json
import logging
import os
from typing import TYPE_CHECKING

import requests

if TYPE_CHECKING:
    from email_parser import Attachment, ParsedEmail

logger = logging.getLogger(__name__)

# Kimi 上下文预算：正文+附件文本合计不超过 6000 字符，留余量给 prompt 和回复
_MAX_CONTENT_CHARS = 6000

_KIMI_API_URL = "https://api.moonshot.cn/v1/chat/completions"
_KIMI_MODEL = "moonshot-v1-8k"


# ---------------------------------------------------------------------------
# 附件文本提取
# ---------------------------------------------------------------------------

def extract_attachment_text(attachment: "Attachment") -> str:
    """
    按 content_type / 文件名后缀分发到对应的提取函数。
    提取失败时返回占位字符串，不抛出异常，让主流程继续。
    """
    ct = attachment.content_type.lower()
    name = attachment.filename.lower()

    try:
        if ct == "application/pdf" or name.endswith(".pdf"):
            return _extract_pdf_text(attachment.data)
        elif ct in (
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            "application/msword",
        ) or name.endswith((".docx", ".doc")):
            return _extract_docx_text(attachment.data)
        elif ct in (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/vnd.ms-excel",
        ) or name.endswith((".xlsx", ".xls")):
            return _extract_xlsx_text(attachment.data)
        elif ct.startswith("image/"):
            return f"[图片附件: {attachment.filename}，POC 阶段暂不支持 OCR]"
        else:
            return f"[不支持的附件类型: {attachment.content_type}，文件名: {attachment.filename}]"
    except Exception as e:
        logger.warning(f"附件 {attachment.filename!r} 文本提取失败: {e}")
        return f"[附件提取失败: {attachment.filename}]"


def _extract_pdf_text(data: bytes) -> str:
    """使用 pdfplumber 提取 PDF 文本（CJK 支持优于 PyPDF2）。"""
    import pdfplumber

    pages_text = []
    with pdfplumber.open(io.BytesIO(data)) as pdf:
        for i, page in enumerate(pdf.pages):
            text = page.extract_text()
            if text:
                pages_text.append(f"[第{i + 1}页]\n{text}")
    return "\n\n".join(pages_text)


def _extract_docx_text(data: bytes) -> str:
    """使用 python-docx 提取 Word 文档文本。"""
    import docx

    doc = docx.Document(io.BytesIO(data))
    return "\n".join(para.text for para in doc.paragraphs if para.text.strip())


def _extract_xlsx_text(data: bytes) -> str:
    """使用 openpyxl 提取 Excel 文本，格式化为逐行输出。"""
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    sheets_text = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(cell) if cell is not None else "" for cell in row]
            if any(c.strip() for c in cells):
                rows.append("\t".join(cells))
        if rows:
            sheets_text.append(f"[Sheet: {sheet_name}]\n" + "\n".join(rows))
    wb.close()
    return "\n\n".join(sheets_text)


# ---------------------------------------------------------------------------
# Kimi LLM 分析
# ---------------------------------------------------------------------------

def _call_kimi(prompt: str, api_key: str) -> str:
    """
    调用 Kimi API。
    完全复用 backend/app/services/ai_service.py 的调用模式。
    """
    response = requests.post(
        _KIMI_API_URL,
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
        },
        json={
            "model": _KIMI_MODEL,
            "messages": [{"role": "user", "content": prompt}],
        },
        timeout=60,
    )
    response.raise_for_status()
    return response.json()["choices"][0]["message"]["content"]


def analyze_email_with_llm(
    parsed_email: "ParsedEmail",
    attachment_texts: dict,
    api_key: str,
) -> dict:
    """
    将邮件正文 + 附件文本送入 Kimi，返回结构化分析结果。

    Args:
        parsed_email: 解析后的邮件对象
        attachment_texts: {filename: extracted_text} 字典
        api_key: Kimi API Key

    Returns:
        dict，包含 summary / category / action_required 等字段
    """
    # 拼接附件文本
    att_sections = []
    for fname, text in attachment_texts.items():
        att_sections.append(f"【附件: {fname}】\n{text}")
    attachment_combined = "\n\n".join(att_sections) if att_sections else "（无附件）"

    # 合并正文 + 附件，截断到预算上限
    body_preview = parsed_email.text_body[:_MAX_CONTENT_CHARS]
    att_preview = attachment_combined[: max(0, _MAX_CONTENT_CHARS - len(body_preview))]

    prompt = f"""你是一个留学申请邮件分析助手。请分析以下邮件内容，以 JSON 格式返回结果。

要求：
- 只返回 JSON，不要任何额外说明
- JSON 字段如下：
  - summary: 一句话摘要（中文）
  - category: 邮件分类，从以下选项选一个：offer / rejection / interview / document_request / notification / other
  - action_required: 用户是否需要操作（true/false）
  - action_description: 如果需要操作，描述具体操作内容；不需要则填 null
  - school_name: 涉及的学校名称（如有，否则填 null）
  - program_name: 涉及的专业/项目名称（如有，否则填 null）
  - deadline: 截止日期（格式 YYYY-MM-DD，如有，否则填 null）

邮件主题: {parsed_email.subject}
发件人: {parsed_email.sender}
发送时间: {parsed_email.date.isoformat() if parsed_email.date else "未知"}

邮件正文:
{body_preview}

{att_preview}"""

    logger.info(f"调用 Kimi 分析邮件 UID={parsed_email.uid}")
    raw_response = _call_kimi(prompt, api_key)

    # 解析 JSON 响应
    try:
        # Kimi 有时会在 JSON 前后加 ```json ... ``` 包裹
        clean = raw_response.strip()
        if clean.startswith("```"):
            clean = clean.split("```")[1]
            if clean.startswith("json"):
                clean = clean[4:]
        result = json.loads(clean.strip())
    except json.JSONDecodeError:
        logger.warning(f"Kimi 返回非 JSON 格式，原始内容: {raw_response[:200]}")
        result = {
            "summary": raw_response[:200],
            "category": "other",
            "action_required": False,
            "action_description": None,
            "school_name": None,
            "program_name": None,
            "deadline": None,
            "_raw": raw_response,
        }

    logger.info(f"UID={parsed_email.uid} LLM 分析结果: {result.get('summary', '')}")
    return result
